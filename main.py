"""Modulo que proporciona una función para obtener"""
from os.path import abspath
from datetime import datetime
from prettytable import PrettyTable
from openpyxl import load_workbook
from matplotlib import pyplot as plt

LOGO = '''
████████╗░█████╗░░██████╗██╗░░██╗  ███╗░░░███╗░█████╗░███╗░░██╗░█████╗░░██████╗░███████╗██████╗░
╚══██╔══╝██╔══██╗██╔════╝██║░██╔╝  ████╗░████║██╔══██╗████╗░██║██╔══██╗██╔════╝░██╔════╝██╔══██╗
░░░██║░░░██ ████║╚█████╗░█████═╝░  ██╔████╔██║██ ████║██╔██╗██║██ ████║██║░░██╗░█████╗░░██████╔╝
░░░██║░░░██╔══██║░╚═══██╗██╔═██╗░  ██║╚██╔╝██║██╔══██║██║╚████║██╔══██║██║░░╚██╗██╔══╝░░██╔══██╗
░░░██║░░░██║░░██║██████╔╝██║░╚██╗  ██║░╚═╝░██║██║░░██║██║░╚███║██║░░██║╚██████╔╝███████╗██║░░██║
░░░╚═╝░░░╚═╝░░╚═╝╚═════╝░╚═╝░░╚═╝  ╚═╝░░░░░╚═╝╚═╝░░╚═╝╚═╝░░╚══╝╚═╝░░╚═╝░╚═════╝░╚══════╝╚═╝░░╚═╝
'''
ABSOLUTA = abspath("tareas.xlsx")
ETIQUETAS = ["Tareas completadas",
                "Tareas No completadas"]
EXP = (0.1, 0.1)
FORMATO_DECIMALES = '%1.1f%%'
wb = load_workbook(ABSOLUTA)
ws = wb.active

COLUMNAS = ['Id. de Tarea', 'Título', 'Fecha de creación', 'Fecha límite', 'Completada']
tabla_tarea = PrettyTable(COLUMNAS)

def actualizar_tabla():
    """Función que refresca la tabla y actualiza los identificadores de las tareas ingresadas"""
    actualizar_identificadores()
    tabla_tarea.clear_rows()
    for fila in ws.iter_rows():
        valores = []
        for celda in fila:
            valores.append(celda.value)
        tabla_tarea.add_row(valores)
    print(tabla_tarea)


def actualizar_identificadores():
    """Funcion que revisa cada identificador de la tabla y lo coloca en orden ascendente"""
    identificador = 1
    for fila in ws.iter_rows(min_row=1, max_col=1):
        for celda in fila:
            celda.value = identificador
            identificador+=1
            wb.save(ABSOLUTA)


def ingrese_fecha_limite():
    """Se solicita ingresar los datos de fecha para
     devolver al usuario la misma fecha en otro formato
     """
    dia_limite = int(input("Ingrese el día de la fecha límite: "))
    mes_limite = int(input("Ingrese el mes de la fecha límite: "))
    anio_limite =  int(input("Ingrese el año de la fecha límite: "))
    return datetime(anio_limite, mes_limite, dia_limite)


def diferencia_dias(fecha_limite, fecha_creacion):
    """Encuentra y muestra la diferencia de días entre 
    la fecha de creación de la tarea y la fecha límite
     """
    dia_creacion = fecha_creacion.date()
    dia_limite_tarea = fecha_limite.date()
    diferencia_de_dias = dia_limite_tarea - dia_creacion
    print(f"Tienes {diferencia_de_dias.days} días para completar la tarea")


def agregar_tarea():
    """Ingresa una nueva tarea dentro del archivo xlsx"""
    fecha_creacion = datetime.now()
    print("Proporcione la fecha límite para realizar esta tarea. ")
    fecha_limite = ingrese_fecha_limite()
    diferencia_dias(fecha_limite, fecha_creacion)
    titulo = input('Inserte el titulo de la tarea: ')
    ws.append(
        (0, 
        titulo,
        fecha_creacion.strftime("%d-%m-%Y"),
        fecha_limite.strftime("%d-%m-%Y"),
        "❌")
        )
    wb.save(ABSOLUTA)


def eliminar_fila(id_tarea):
    """Elimina nuevas tareas en cada valor de """
    for fila in ws.iter_rows(min_row=1, max_col=1):
        for celda in fila:
            if celda.value == id_tarea:
                ws.delete_rows(celda.row)
                wb.save(ABSOLUTA)
                return True
    return False


def editar_fila(id_tarea, edicion_titulo, edicion_fecha, tarea_completada):
    """Pregunta al usuario si quiere modificar cada elemento de la tarea,
     para luego solicitarle que ingrese los valores.
     """
    if id_tarea <= ws.max_row and id_tarea > 0:
        fila = ws[id]
        if edicion_titulo == 's':
            fila[1].value = input("Ingrese el titulo de la tarea: ")
        if edicion_fecha == 's':
            fila[3].value = ingrese_fecha_limite().strftime("%d-%m-%Y")
        if tarea_completada == 's':
            fila[4].value = '✅'
        else:
            fila[4].value = '❌'
        wb.save(ABSOLUTA)
    else:
        print("Ingrese un identificador valido.")


def mostrar_estadistica():
    """Muestra en una gráfica que aparece en una ventana emergente, mostrando el porcentaje
       de tareas completadas 
    """
    tareas_totales = ws.max_row
    tareas_completadas = 0
    for fila in ws.iter_rows(min_row=1, min_col=5, max_col=5):
        for celda in fila:
            if celda.value == '✅':
                tareas_completadas+=1
    tareas_faltantes = tareas_totales-tareas_completadas
    tareas_completadas/=tareas_totales*100
    tareas_faltantes/=tareas_totales*100
    valores = [tareas_completadas, tareas_faltantes]
    
    plt.pie(valores, explode=EXP, labels=ETIQUETAS, autopct=FORMATO_DECIMALES, shadow=True, startangle=90)
    plt.title("Estadísticas de tareas")
    plt.legend()
    plt.show()

print(LOGO)
actualizar_tabla()
print('')

respuesta_usuario = input('¿Qué desea hacer? a: Agregar tarea | b: Eliminar | \
c: Editar | e: Mostrar estadística | s: Salir: ')

while respuesta_usuario != 's':

    if respuesta_usuario == 'a':
        agregar_tarea()

    elif respuesta_usuario == 'b':

        dato_eliminar = int(input("Ingrese el Id. de la tarea a eliminar: "))
        if eliminar_fila(dato_eliminar):
            print("Tarea eliminada exitosamente :)")
        else:
            print("No se ha podido eliminar la fila :(")

    elif respuesta_usuario == 'c':
        dato_editar = int(input("Ingrese el Id. del dato que desea modificar: "))
        editar_titulo = input('¿Desea editar el titulo? s/n: ')
        editar_fecha = input('¿Desea editar la fecha limite? s/n: ')
        completada = input('¿Ya completaste la tarea? s/n: ')
        editar_fila(dato_editar, editar_titulo, editar_fecha, completada)
    elif respuesta_usuario=='e':
        mostrar_estadistica()
    else:
        print("Seleccione una opción válida.")

    actualizar_tabla()
    respuesta_usuario = input('¿Qué desea hacer? a: Agregar tarea | b: Eliminar | \
    c: Editar | e: Mostrar estadística | s: Salir: ')


wb.close()

print('\x1b[6;30;42m' + 'Gracias por usar el programa!' + '\x1b[0m')
# End-of-file